"""Add/refresh the 'Powers' and 'Power List' sheets in hero-stats.xlsx.

Naming tiers, per design:
  0,1,2  shared by every hero of the same PRIMARY type (rank 0 = the auto-attack)
  3      shared by every hero of the same SECONDARY type
  4,5    unique to the hero

Only the six arcane types ever appear as a secondary — melee heroes always take
an arcane secondary, so there are six tier-3 powers, not nine.

'Powers' is the per-hero grid; 'Power List' is one row per *distinct* name, with
who owns it and a wide Prompt column to describe what it does.

Run:  python tools/add-powers-sheet.py
Rewrites both sheets; Hero Stats is left untouched. **Prompts already written on
Power List are carried across**, matched by power name — rename a power and its
prompt is orphaned (the script says so rather than dropping it silently).
"""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# --- tier 1: by primary type. Voice follows each House in LORE-and-flavor.md ---
BY_PRIMARY = {
    "Earth":  ["Sundered Ground", "Root and Hold", "The Mountain Answers"],
    "Air":    ["Cutting Gale", "Thin the Air", "The Sky Unmoored"],
    "Fire":   ["Ember Lash", "Feed the Bloom", "Everything Burns Twice"],
    "Water":  ["Undertow", "Wear the Stone", "The Tide Remembers"],
    "Light":  ["Verdict Light", "Nothing Hidden", "The First Word Spoken"],
    "Dark":   ["Veilcut", "Draw the Veil", "The Last Silence"],
    # "Redouble" is the fencing term for a renewed attack of your own, not a
    # counter — "Riposte" read as a reactive power once reactions became real.
    "Slash":  ["Open Line", "Redouble", "One Clean Stroke"],
    "Pierce": ["Seamfinder", "Thread the Guard", "The Single Truth"],
    "Crush":  ["Deadweight", "Make an Opening", "The Sky Falls"],
}

# --- tier 2: by secondary type. Reads as the second Force lending itself ---
BY_SECONDARY = {
    "Earth": "The Deep Lends Weight",
    "Air":   "The Sky Lends Swiftness",
    "Fire":  "The Bloom Lends Heat",
    "Water": "The Tide Lends Patience",
    "Light": "The Word Lends Sight",
    "Dark":  "The Silence Lends Cover",
}

# --- tier 3: unique, drawn from each hero's concept and epithet ---
UNIQUE = {
    "Bramwen":           ("Years in the Making", "Wrath of the Slow Stone"),
    "Ossic":             ("Kneel and Raise", "The God-Bone Wakes"),
    "Terragosa":         ("Orchard Over Ruin", "The Green Crown Descends"),
    "Zephyrine":         ("A Distance Never Closed", "The Thin Blade Falls"),
    # The three Buffers' P4 slots carry the game's only healing — one each at
    # single-target / row / party scale. They replaced a redundant second
    # buff-strip or speed-buff apiece.
    "Cirrolan":          ("Fair Weather", "Whisper from the High Reach"),
    "Vael":              ("Jump First", "Nothing to Catch You"),
    "Ember Saelith":     ("The Room Warms", "First Spark, Last Laugh"),
    "Pyrrhic":           ("Less Left to Lose", "Glad Ruin"),
    "Cindara":           ("The Coal, Not the Flame", "The Long Burn"),
    "Marisel":           ("Your Own Past, Rising", "Drown in What You Did"),
    "Tidewarden Coll":   ("Give Ground, Take Coast", "The Bulwark Holds"),
    "Nix":               ("Perfectly Calm", "The Still Pool Closes"),
    "Seraphel":          ("The Gaze Accuses", "Sentence Passed"),
    "Lucen":             ("Enough Light for Everyone", "The Unhidden Hour"),
    "Auriel Dawnkeep":   ("The Lantern Holds", "Last Light on the Wall"),
    "Nyxara":            ("A Kindness, Ending", "Mercy at the End"),
    "Umbriel":           ("Unmake the Wound", "The Undoing"),
    "Corvane":           ("I Know Your Hour", "Shepherd of Endings"),
    "Kaellis":           ("Never Twice", "Immaculate"),
    "Reyna Two-Rivers":  ("Two Rivers Meeting", "The Current Takes All"),
    "Grieve":            ("Clear the Room", "The Wide Reaping"),
    "Vantric":           ("The One Gap", "The Spear Finds It"),
    "Silka Pinquick":    ("Already Behind You", "Quicker Than Told"),
    "Lord Aiguille":     ("Before You Decide", "The Long Point"),
    "Boldrek":           ("All At Once", "Avalanche"),
    "Hettamar Ironfall": ("End of Argument", "Ironfall"),
    "Mauless":           ("Guards Break First", "The Undenied"),
}

INK = "FFF4EFE4"
HEAD_FILL = PatternFill("solid", fgColor="FF241F38")
PRIMARY_FILL = PatternFill("solid", fgColor="FFEFF4EC")    # tiers share a tint
SECONDARY_FILL = PatternFill("solid", fgColor="FFF2F0F7")
UNIQUE_FILL = PatternFill("solid", fgColor="FFFFFDF5")
AUTO_FILL = PatternFill("solid", fgColor="FFE6EEE2")

HEAD = ["#", "Role", "Hero", "Primary", "Secondary",
        "Power 0 — auto", "Power 1", "Power 2", "Power 3", "Power 4", "Power 5"]

LIST_HEAD = ["Power", "Tier", "Shared via", "Elements", "Heroes", "Prompt — what does it do?"]

TIER_FILL = {0: AUTO_FILL, 1: PRIMARY_FILL, 2: PRIMARY_FILL,
             3: SECONDARY_FILL, 4: UNIQUE_FILL, 5: UNIQUE_FILL}


def read_heroes(src):
    """Pull the roster out of Hero Stats, in sheet order."""
    head = [c.value for c in src[1]]
    col = {h: head.index(h) + 1 for h in head if h}
    heroes = []
    for r in range(2, src.max_row + 1):
        name = src.cell(r, col["Hero"]).value
        if not name:
            continue
        heroes.append({
            "n": src.cell(r, col["#"]).value,
            "name": name,
            "role": src.cell(r, col["Role"]).value,
            "pri": src.cell(r, col["Primary"]).value,
            "sec": src.cell(r, col["Secondary"]).value,
        })
    return heroes


def existing_prompts(wb):
    """Carry authored prompts across a rebuild, keyed by power name."""
    if "Power List" not in wb.sheetnames:
        return {}
    ws = wb["Power List"]
    head = [c.value for c in ws[1]]
    if "Power" not in head or LIST_HEAD[-1] not in head:
        return {}
    kc, pc = head.index("Power") + 1, head.index(LIST_HEAD[-1]) + 1
    kept = {}
    for r in range(2, ws.max_row + 1):
        key, val = ws.cell(r, kc).value, ws.cell(r, pc).value
        if key and val and str(val).strip():
            kept[key] = val
    return kept


def head_row(ws, labels, height=30):
    ws.append(labels)
    for c in range(1, len(labels) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color=INK)
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = height


def write_legend(ws, col, lines):
    for i, (text, bold) in enumerate(lines, start=1):
        cell = ws.cell(i, col, text)
        if bold:
            cell.font = Font(bold=True, color=INK)
            cell.fill = HEAD_FILL
        else:
            cell.font = Font(size=9)
    ws.column_dimensions[get_column_letter(col)].width = 66


def build_powers(wb, heroes):
    """The per-hero grid: one row per hero, six named powers across."""
    if "Powers" in wb.sheetnames:
        del wb["Powers"]
    ws = wb.create_sheet("Powers", wb.sheetnames.index("Hero Stats") + 1)
    head_row(ws, HEAD)

    for i, h in enumerate(heroes):
        r = i + 2
        row = [h["n"], h["role"], h["name"], h["pri"], h["sec"],
               *BY_PRIMARY[h["pri"]], BY_SECONDARY[h["sec"]], *UNIQUE[h["name"]]]
        for c, v in enumerate(row, start=1):
            ws.cell(r, c, v)
        for c in range(6, 12):
            ws.cell(r, c).fill = TIER_FILL[c - 6]
        ws.cell(r, 6).font = Font(italic=True)

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:K{len(heroes) + 1}"
    for c, w in {"A": 5, "B": 10, "C": 20, "D": 11, "E": 12, "F": 22,
                 "G": 22, "H": 26, "I": 26, "J": 26, "K": 28}.items():
        ws.column_dimensions[c].width = w

    write_legend(ws, 13, [
        ("HOW POWERS ARE SHARED", True),
        ("Power 0 — auto-skill. The default attack, usable every turn with", False),
        ("no cooldown, so a hero is never idle. Shared by primary type.", False),
        ("Powers 1-2 — shared by every hero of the same PRIMARY type.", False),
        ("9 types x 3 = 27 names. Three champions of a House share them.", False),
        ("Power 3 — shared by every hero of the same SECONDARY type.", False),
        ("Only 6 names: melee always takes an arcane secondary, so no", False),
        ("melee type ever appears here.", False),
        ("Powers 4-5 — unique to the hero. 54 names, the identity layer.", False),
        ("", False),
        ("Effects, costs and cooldowns are drafted on the Power List sheet —", False),
        ("see mechanics/03-powers.md when it exists.", False),
    ])
    return ws


def collect_distinct(heroes):
    """One entry per distinct power name, in tier then roster order."""
    rows, seen = [], {}

    def add(name, tier, via, elements, hero):
        if name in seen:
            seen[name]["heroes"].append(hero)
            return
        seen[name] = {"power": name, "tier": tier, "via": via,
                      "elements": elements, "heroes": [hero]}
        rows.append(seen[name])

    for tier in range(3):
        for h in heroes:
            add(BY_PRIMARY[h["pri"]][tier], tier, "Primary", h["pri"], h)
    for h in heroes:
        add(BY_SECONDARY[h["sec"]], 3, "Secondary", h["sec"], h)
    for tier in (4, 5):
        for h in heroes:
            add(UNIQUE[h["name"]][tier - 4], tier, "Unique",
                f'{h["pri"]} · {h["sec"]}', h)

    rows.sort(key=lambda e: (e["tier"], min(x["n"] for x in e["heroes"])))
    return rows


def build_power_list(wb, heroes, carried):
    """One row per distinct name, with owners and a wide Prompt column."""
    if "Power List" in wb.sheetnames:
        del wb["Power List"]
    ws = wb.create_sheet("Power List", wb.sheetnames.index("Powers") + 1)
    head_row(ws, LIST_HEAD)

    rows = collect_distinct(heroes)
    for i, e in enumerate(rows):
        r = i + 2
        owners = " · ".join(f'{h["name"]} ({h["role"]})' for h in e["heroes"])
        tier = "0 — auto" if e["tier"] == 0 else e["tier"]
        for c, v in enumerate([e["power"], tier, e["via"], e["elements"],
                               owners, carried.get(e["power"])], start=1):
            ws.cell(r, c, v)
        ws.cell(r, 1).font = Font(bold=True)
        ws.cell(r, 1).fill = TIER_FILL[e["tier"]]
        for c in (5, 6):
            ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
        for c in (2, 3):
            ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="top")

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:F{len(rows) + 1}"
    for c, w in {"A": 28, "B": 9, "C": 11, "D": 18, "E": 46, "F": 90}.items():
        ws.column_dimensions[c].width = w

    write_legend(ws, 8, [
        ("ONE ROW PER DISTINCT POWER", True),
        ("87 names: 27 shared by primary type (tiers 0-2), 6 shared by", False),
        ("secondary type (tier 3), 54 unique to a hero (tiers 4-5).", False),
        ("", False),
        ("Elements — the type that grants the power. For a unique power", False),
        ("that is the owner's primary and secondary both.", False),
        ("Heroes — every hero that gets this power, with its role, since", False),
        ("a shared power has to read well for all of them at once.", False),
        ("", False),
        ("Prompt — free text: what the power should do. Prompts survive a", False),
        ("rerun of tools/add-powers-sheet.py, matched on the power name.", False),
        ("Rename a power and its prompt is orphaned, not moved.", False),
    ])
    return rows


def main() -> None:
    path = Path(__file__).resolve().parent.parent / "resources" / "characters" / "hero-stats.xlsx"
    wb = load_workbook(path)
    heroes = read_heroes(wb["Hero Stats"])
    carried = existing_prompts(wb)

    build_powers(wb, heroes)
    rows = build_power_list(wb, heroes, carried)
    wb.save(path)

    names = {e["power"] for e in rows}
    orphans = sorted(set(carried) - names)
    by_tier = {t: sum(1 for e in rows if e["tier"] == t) for t in range(6)}
    print(f"Powers sheet:     {len(heroes)} heroes x 6 powers, Role in column B")
    print(f"Power List sheet: {len(rows)} distinct names "
          f"({'/'.join(str(by_tier[t]) for t in range(6))} by tier 0-5)")
    print(f"prompts carried across: {len(carried) - len(orphans)}")
    if orphans:
        print(f"ORPHANED prompts (power renamed or removed): {', '.join(orphans)}")


if __name__ == "__main__":
    main()
