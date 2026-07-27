"""Apply the settled power-balance pass to resources/characters/hero-stats.xlsx.

Edits the workbook in place. Prompts already authored are preserved except
where a decision explicitly rewrites one.

What this applies, all decided in conversation:

  * multiplier curve smoothed to 1.0 / 1.5 / 2.0 / 2.5 / 3.5 / 5.0, so every
    tier is strictly better than the one below instead of three flat bands
  * cooldowns 0/1/2/3/5/7 by tier, with the 54 unique powers varying per hero
    off the owner's Speed band - fast heroes cycle, heavy hitters wait
  * tiers 4-5 gated: T4 unavailable until turn 3, T5 until turn 5
  * multi-target powers resolve PER TARGET at a reduced multiplier
  * the three Buffer P4 slots become the game's only healing
  * "Riposte" renamed - it read as a reactive power once reactions became real
  * taunt / fade named explicitly where prompts said "draws attention" or
    "strips stealth"
  * Marisel's Reckoning stacks given a single defined source

Run:  python tools/apply-power-balance.py
Idempotent - safe to re-run.
"""

import re
from pathlib import Path

from openpyxl import load_workbook


def sync_numbers(text, mult, cd):
    """Rewrite the numbers *stated in a prompt* to match its columns.

    The prompts open with phrases like "Might x 1.75 single-target Earth strike,
    1-turn cooldown". Changing the curve silently falsified 69 of them, and a
    prompt that disagrees with its own row is worse than no prompt at all.
    """
    if not text:
        return text, False
    before = text
    if mult is not None:
        pretty = f"{mult:g}"
        text = re.sub(r"(?i)\bMight\s*[x×]\s*[\d.]+", f"Might x{pretty}", text)
    if cd is not None and cd > 0:
        text = re.sub(r"\b\d+-turn cooldown", f"{cd:g}-turn cooldown", text)
    return text, text != before

TIER_MULT = {"0 — auto": 1.0, "1": 1.5, "2": 2.0, "3": 2.5, "4": 3.5, "5": 5.0}
TIER_CD = {"0 — auto": 0, "1": 1, "2": 2, "3": 3, "4": 5, "5": 7}

# Unique cooldowns vary off the owner's Speed band. Faster heroes cycle their
# signature powers sooner; this is the cheapest lever that makes 27 heroes play
# at different tempos without touching a single multiplier.
SPEED_BAND_CD = {45: (4, 6), 35: (4, 6), 30: (5, 7), 25: (5, 7), 15: (6, 8)}

RENAMES = {
    "Riposte": "Redouble",
    "Rumour and Storm": "Fair Weather",
    "Three Beams, No Shadow": "Enough Light for Everyone",
    "Unwrite the Line": "Unmake the Wound",
}

# Multi-target powers: per target, at a reduced multiplier. Total output lands
# ~1.4x a single-target power of the same tier, so breadth costs per-head power.
AOE_MULT = {
    "Two Rivers Meeting": 2.5,      # 2 targets, tier 4
    "Clear the Room": 1.75,         # one row (up to 3), tier 4
    "The Current Takes All": 2.5,   # 3 targets, tier 5
    "The Wide Reaping": 2.5,        # one row, tier 5
}

# Healing scales off Might like everything else - Might is deliberately
# type-agnostic and is the only offense stat, so a heal has no other anchor.
HEAL_MULT = {
    "Unmake the Wound": 3.5,            # strong, single target
    "Fair Weather": 1.75,               # moderate, one row
    "Enough Light for Everyone": 1.0,   # light, whole party
}

# Neither damage nor healing -> skips phase 3 entirely, so no multiplier exists
# to record. Blank, not zero: zero would read as "deals no damage" rather than
# "damage is not a thing this power has".
NO_NUMBER = ["Whisper from the High Reach", "The Unhidden Hour", "The Undoing"]

PROMPTS = {
    "Redouble":
        "Might x1.5 single-target Slash strike, 1-turn cooldown, no rider — a clean, "
        "unembellished second attack. Flavor: the duelist's redouble, a renewed thrust "
        "after the first is answered. (Renamed from 'Riposte': that read as a reactive "
        "power once reactions became a real mechanic, and this one fires on your own turn.)",
    "Unmake the Wound":
        "STRONG SINGLE-TARGET HEAL. Restores Might x3.5 to one ally, no damage component. "
        "Runs the Defense phase like an attack — healing is the same operation with the "
        "sign reversed — and is reach-limited exactly as an attack is. Flavor: Umbriel "
        "unwrites the wound rather than closing it; what she takes back was never there.",
    "Fair Weather":
        "MODERATE ROW HEAL. Restores Might x1.75 to every ally in one row, per target, no "
        "damage component. Reach-limited like any other power. Flavor: the wind turns "
        "kind for a moment, and a whole line breathes easier.",
    "Enough Light for Everyone":
        "LIGHT PARTY HEAL. Restores Might x1 to every ally in the squad, per target, no "
        "damage component. The smallest per-head heal in the game and the only one that "
        "reaches all six. Flavor: not much each, but nobody is left standing in the dark.",
    "Your Own Past, Rising":
        "Might x3.5 single-target Water/Dark strike, tier 4. Damage scales with the "
        "target's RECKONING stacks and does NOT consume them. Stacks come from one place "
        "only — Marisel's passive 'It All Comes Back' adds one each time she damages that "
        "target — so this power reads them, it does not build them. Flavor: she drowns "
        "you in your own mistakes, one resurfacing at a time.",
    "Drown in What You Did":
        "Might x5 single-target Water/Dark strike, tier 5. CONSUMES every Reckoning stack "
        "on the target for a finishing burst, scaling with how many were spent. Her whole "
        "kit is one mechanic: the passive banks, tier 4 reads, tier 5 spends. Flavor: "
        "every past hit comes due at once.",
    "The Bulwark Holds":
        "Might x5 single-target Water/Earth strike, grants a larger Armor buff, and "
        "applies TAUNT to Coll for a short duration — enemies that can reach him must "
        "target him. Taunt narrows a candidate set and never extends one, so it binds "
        "only attackers already in range of him. Flavor: the bulwark holds long after "
        "anyone expected it to.",
    "Last Light on the Wall":
        "Might x5 single-target Light/Water strike, grants a larger Resolve and Toughness "
        "buff, and applies TAUNT to Auriel for a short duration. Note her Tank role "
        "passive already taunts within her own row; this extends the compulsion to every "
        "enemy that can reach her. Flavor: the last light on the wall will not go out.",
    "Nothing Hidden":
        "Might x1.5 single-target Light strike, 1-turn cooldown. Rider: removes FADE from "
        "the target, so it can be selected normally even while unfaded allies stand. "
        "Flavor: whatever the target hoped would stay unseen, doesn't anymore.",
    "The First Word Spoken":
        "Might x2 single-target Light strike, 2-turn cooldown. Rider: removes FADE from "
        "the target and lowers its accuracy — being fully seen throws off more than just "
        "cover. Flavor: the first word is spoken again, and this time it costs something.",
}


def main() -> None:
    path = Path(__file__).resolve().parent.parent / "resources" / "characters" / "hero-stats.xlsx"
    wb = load_workbook(path)

    hs = wb["Hero Stats"]
    head = [c.value for c in hs[1]]
    hc = {h: head.index(h) + 1 for h in head if h}
    speed = {}
    for r in range(2, hs.max_row + 1):
        n = hs.cell(r, hc["Hero"]).value
        if n:
            speed[n] = hs.cell(r, hc["Speed"]).value

    # --- renames on the per-hero grid ---
    pw = wb["Powers"]
    renamed = 0
    for r in range(2, pw.max_row + 1):
        for c in range(6, 12):
            v = pw.cell(r, c).value
            if v in RENAMES:
                pw.cell(r, c, RENAMES[v])
                renamed += 1

    # --- the distinct-power list ---
    pl = wb["Power List"]
    ph = [pl.cell(1, c).value for c in range(1, 9)]
    C = {h: ph.index(h) + 1 for h in ph if h}
    NAME, TIER, MULT, CD, PROMPT = (C["Power"], C["Tier"], C["Power Multiplier"],
                                    C["Cooldown"], C["Prompt — what does it do?"])
    HEROES = C["Heroes"]

    stats = {"renamed": renamed, "mult": 0, "cd": 0, "prompt": 0, "blank": 0, "synced": 0}
    for r in range(2, pl.max_row + 1):
        name = pl.cell(r, NAME).value
        if not name:
            continue
        if name in RENAMES:
            name = RENAMES[name]
            pl.cell(r, NAME, name)
        tier = str(pl.cell(r, TIER).value or "").strip()
        if tier == "passive":
            continue

        # multiplier: tier curve, then AoE and heal overrides, then blanks
        if name in NO_NUMBER:
            pl.cell(r, MULT, None)
            stats["blank"] += 1
        else:
            m = AOE_MULT.get(name) or HEAL_MULT.get(name) or TIER_MULT.get(tier)
            if m is not None:
                pl.cell(r, MULT, m)
                stats["mult"] += 1

        # cooldown: tier default, or the owner's Speed band for a unique
        cd = TIER_CD.get(tier)
        if tier in ("4", "5"):
            owner = str(pl.cell(r, HEROES).value or "").split(" (")[0]
            band = SPEED_BAND_CD.get(speed.get(owner))
            if band:
                cd = band[0] if tier == "4" else band[1]
            if name in AOE_MULT:      # breadth costs an extra turn of downtime
                cd += 1
        if cd is not None:
            pl.cell(r, CD, cd)
            stats["cd"] += 1

        if name in PROMPTS:
            pl.cell(r, PROMPT, PROMPTS[name])
            stats["prompt"] += 1

        # last: make every prompt's stated numbers agree with its own row
        synced, changed = sync_numbers(pl.cell(r, PROMPT).value,
                                       pl.cell(r, MULT).value, pl.cell(r, CD).value)
        if changed:
            pl.cell(r, PROMPT, synced)
            stats["synced"] += 1

    wb.save(path)
    for k, v in stats.items():
        print(f"  {k:<8} {v}")
    print("\ngating (not a cell — a rule for 03-powers.md):")
    print("  tier 4 unavailable until turn 3 · tier 5 until turn 5")


if __name__ == "__main__":
    main()
