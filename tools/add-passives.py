"""Add the passive layer to resources/characters/hero-stats.xlsx.

Every hero gets THREE passives, one per sharing scope, and each scope is a
different *kind* of effect so they compose instead of stacking:

  Role   (4)  positional and tempo - rows, reach, durations
  House  (9)  the Force's mechanical signature - what the damage type does
  Unique (27) a conditional trigger, never a flat number

40 distinct passives. Appends to both sheets rather than rebuilding them, so
the multipliers, cooldowns and prompts already authored are left untouched.

Run:  python tools/add-passives.py
Idempotent: re-running removes the previously added passive rows/columns first.
"""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# --- Role: how the hero relates to rows, reach and the duration clock -------
BY_ROLE = {
    "Striker": ("Finish It",
                "Bonus damage against any target already below half its pool. Strikers "
                "carry the highest Might on the board and hold the two front slots, so "
                "their passive rewards closing out rather than opening."),
    "Tank":    ("Hold the Line",
                "Allies sharing this hero's row take reduced damage while it is alive. "
                "Positional rather than numeric - it pays for putting the tank where the "
                "formation needs it, and it stops mattering the moment the tank falls."),
    "Ranged":  ("Measured Shot",
                "Bonus damage against targets at the far edge of this hero's reach "
                "(distance 2). Rewards staying back and punishes an enemy that closes, "
                "which is the whole reason Ranged pays a lower stat budget."),
    "Buffer":  ("Long Counsel",
                "Buffs and heals this hero applies last one turn longer. Ties straight "
                "to the duration clock in 04-turns.md - durations tick on the bearer's "
                "own turn, so this is worth more on a slow ally than a fast one."),
}

# --- House: the Force's signature, one per primary type ---------------------
BY_HOUSE = {
    "Earth":  ("The Deep Holds",
               "Control effects on this hero expire one turn sooner. The Rooted Deep "
               "does not stay held."),
    "Air":    ("Never Where You Struck",
               "Whenever an attack misses this hero, it gains Agility until its next "
               "turn. Missing an Air champion makes the next attempt worse."),
    "Fire":   ("It Catches",
               "Burning effects this hero applies grow with each tick rather than "
               "dealing a flat amount. Fire escalates; it does not hold steady."),
    "Water":  ("Wears Through",
               "Mitigation reductions this hero applies last one turn longer. Water "
               "does not hit harder - it keeps the seam open."),
    "Light":  ("Nothing Stays Hidden",
               "This hero ignores fade: a faded enemy is targetable normally, without "
               "waiting for the unfaded heroes in front of it. The one House that reads "
               "straight into the targeting pipeline."),
    "Dark":   ("The Veil Closes",
               "Whenever an enemy within this hero's reach is defeated, it gains a brief "
               "buff. The Last Silence feeds on endings, whoever caused them."),
    "Slash":  ("The Cut Reopens",
               "This hero's critical hits apply a bleed. A clean cut is the one that "
               "does not close."),
    "Pierce": ("Find the Seam",
               "This hero's Penetration rises against any target it has already struck "
               "this battle. The first thrust is a question; the second knows."),
    "Crush":  ("Nothing Holds",
               "This hero's attacks shave the target's Armor, and it stacks. Crush does "
               "not go through the guard - it removes the guard."),
}

# --- Unique: one conditional trigger per hero ------------------------------
UNIQUE = {
    "Bramwen":           ("The Long Patience", "Gains a small permanent Might increase at the end of each of her own turns, for the rest of the battle. The longer she is left standing, the worse the eventual answer."),
    "Ossic":             ("The Bone Beneath", "Gains Armor while below half his pool. The god-bone surfaces only when he is pressed onto it."),
    "Terragosa":         ("Something Green Returns", "Recovers a small amount of pool at the start of each of her own turns. Not a heal she casts - a thing that simply keeps growing back."),
    "Zephyrine":         ("Out of Reach", "Takes reduced damage from attackers with reach 1. The distance she keeps is not a position, it is a habit."),
    "Cirrolan":          ("Word Travels", "Buffs he applies also reach one additional ally in the same row as the target. Nothing he says stays where he said it."),
    "Vael":              ("Gravity Is a Suggestion", "Acts first in the opening round regardless of Speed. He jumped before the order was called."),
    "Ember Saelith":     ("Never Quite Out", "Burning effects she applies cannot be cleansed. They can expire; they cannot be put out."),
    "Pyrrhic":           ("Nothing Left to Take", "His damage rises as his own pool falls. The passive version of his whole kit, and the reason he is built to be hurt."),
    "Cindara":           ("Banked Coals", "Damage-over-time effects she applies last one turn longer. She is the coal, not the flame."),
    "Marisel":           ("It All Comes Back", "Bonus damage against any target she has already struck this battle. Feeds Your Own Past, Rising without needing it."),
    "Tidewarden Coll":   ("Ground Yielded", "Takes reduced damage while any ally occupies a row in front of him. A bulwark that is only a bulwark with something to hold."),
    "Nix":               ("No Ripple", "Immune to accuracy reduction. Nothing disturbs the surface enough to spoil the aim."),
    "Seraphel":          ("Under Judgement", "Bonus damage against targets carrying no buffs. She hits hardest once every excuse is gone."),
    "Lucen":             ("Nothing Casts Twice", "His heals also remove one debuff from the target. Light does not clean up afterwards; it arrives already clean."),
    "Auriel Dawnkeep":   ("Still Burning", "Once per battle, survives an otherwise-fatal blow at 1 pool. The lantern goes out on the second hit, not the first."),
    "Nyxara":            ("Merciful", "Restores a small amount of her own pool whenever she defeats a target. An ending is a kindness both ways."),
    "Umbriel":           ("Written in Pencil", "Debuffs she applies cannot be cleansed. What she unwrites stays unwritten."),
    "Corvane":           ("The Ledger Kept", "Gains Resolve each time any hero on the field is defeated, either side. He has been expecting all of them."),
    "Kaellis":           ("The Duelist's Habit", "Bonus damage against any target he has not yet struck this battle - the exact inverse of Marisel. He was always best on the opening exchange."),
    "Reyna Two-Rivers":  ("Confluence", "Her multi-target attacks gain damage for each additional target caught. Two currents meeting is more than either one."),
    "Grieve":            ("Room to Swing", "Gains Armor for each enemy currently within his reach. Surrounded is where the scythe wants to be."),
    "Vantric":           ("Seams Everywhere", "Ignores a flat amount of the target's mitigation on every attack, before Penetration is applied at all."),
    "Silka Pinquick":    ("Already Gone", "Cannot be the target of reactive powers. By the time the counter comes, she is not standing there."),
    "Lord Aiguille":     ("First Guard", "The first attack against him each battle is reduced. His guard is up before the battle is."),
    "Boldrek":           ("No Warning", "His critical hits deal additional damage on top of the crit. An avalanche does not build up."),
    "Hettamar Ironfall": ("Nothing to Discuss", "Any enemy he damages cannot fire a reactive power that turn. The argument ends when he says it does."),
    "Mauless":           ("Immovable", "Cannot be compelled by taunt - he always chooses his own target. Nothing gets to decide where he swings."),
}

INK = "FFF4EFE4"
HEAD_FILL = PatternFill("solid", fgColor="FF241F38")
ROLE_FILL = PatternFill("solid", fgColor="FFEDF1F6")
HOUSE_FILL = PatternFill("solid", fgColor="FFF6F0EC")
UNIQ_FILL = PatternFill("solid", fgColor="FFFFFDF5")

PASSIVE_HEAD = ["Passive · Role", "Passive · House", "Passive · Unique"]
SCOPE_FILL = {"Role": ROLE_FILL, "House": HOUSE_FILL, "Unique": UNIQ_FILL}


def roster(wb):
    src = wb["Hero Stats"]
    head = [c.value for c in src[1]]
    col = {h: head.index(h) + 1 for h in head if h}
    out = []
    for r in range(2, src.max_row + 1):
        name = src.cell(r, col["Hero"]).value
        if not name:
            continue
        out.append({"n": src.cell(r, col["#"]).value, "name": name,
                    "role": src.cell(r, col["Role"]).value,
                    "pri": src.cell(r, col["Primary"]).value,
                    "sec": src.cell(r, col["Secondary"]).value})
    return out


def last_row(ws, col=1):
    return max(r for r in range(2, ws.max_row + 1) if ws.cell(r, col).value)


def add_powers_columns(wb, heroes):
    """Three passive columns on the per-hero grid, after Power 5."""
    ws = wb["Powers"]
    head = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if PASSIVE_HEAD[0] in head:                      # idempotent
        first = head.index(PASSIVE_HEAD[0]) + 1
        ws.delete_cols(first, 3)
    at = 12
    ws.insert_cols(at, 3)
    for i, label in enumerate(PASSIVE_HEAD):
        cell = ws.cell(1, at + i, label)
        cell.font = Font(bold=True, color=INK)
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, h in enumerate(heroes):
        r = i + 2
        vals = (BY_ROLE[h["role"]][0], BY_HOUSE[h["pri"]][0], UNIQUE[h["name"]][0])
        fills = (ROLE_FILL, HOUSE_FILL, UNIQ_FILL)
        for j, (v, f) in enumerate(zip(vals, fills)):
            c = ws.cell(r, at + j, v)
            c.fill = f
    for j, w in enumerate((22, 24, 26)):
        ws.column_dimensions[ws.cell(1, at + j).column_letter].width = w
    ws.auto_filter.ref = f"A1:N{len(heroes) + 1}"
    return ws


def add_power_list_rows(wb, heroes):
    """40 rows appended to the distinct-power list."""
    ws = wb["Power List"]
    start = last_row(ws) + 1
    # idempotent: strip any previously added passive rows
    for r in range(ws.max_row, 1, -1):
        if str(ws.cell(r, 2).value or "").strip().lower() == "passive":
            ws.delete_rows(r)
    start = last_row(ws) + 1

    rows = []
    for role, (name, prompt) in BY_ROLE.items():
        owners = [h for h in heroes if h["role"] == role]
        rows.append((name, "Role", role, owners, prompt))
    for house, (name, prompt) in BY_HOUSE.items():
        owners = [h for h in heroes if h["pri"] == house]
        rows.append((name, "House", house, owners, prompt))
    for h in heroes:
        name, prompt = UNIQUE[h["name"]]
        rows.append((name, "Unique", f'{h["pri"]} · {h["sec"]}', [h], prompt))

    for i, (name, scope, elements, owners, prompt) in enumerate(rows):
        r = start + i
        who = " · ".join(f'{o["name"]} ({o["role"]})' for o in owners)
        for c, v in enumerate([name, "passive", scope, elements, who, None, None, prompt], start=1):
            ws.cell(r, c, v)
        ws.cell(r, 1).font = Font(bold=True)
        ws.cell(r, 1).fill = SCOPE_FILL[scope]
        for c in (5, 8):
            ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
        for c in (2, 3):
            ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="top")
    ws.auto_filter.ref = f"A1:H{start + len(rows) - 1}"
    return rows


def main() -> None:
    path = Path(__file__).resolve().parent.parent / "resources" / "characters" / "hero-stats.xlsx"
    wb = load_workbook(path)
    heroes = roster(wb)
    add_powers_columns(wb, heroes)
    rows = add_power_list_rows(wb, heroes)
    wb.save(path)

    names = {r[0] for r in rows}
    assert len(names) == 40, f"expected 40 distinct passives, got {len(names)}"
    print(f"Powers sheet:     3 passive columns added for {len(heroes)} heroes")
    print(f"Power List sheet: {len(rows)} passive rows appended "
          f"({len(BY_ROLE)} role + {len(BY_HOUSE)} house + {len(UNIQUE)} unique)")
    print(f"every hero now has 6 active powers + 3 passives = 9")


if __name__ == "__main__":
    main()
