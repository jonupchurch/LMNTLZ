"""Generate resources/characters/hero-stats.xlsx — the hero base-stat worksheet.

Bane and Fault are COMPUTED from the counter bijection, never hand-authored,
per the derivation rule in CLAUDE.md. Roster data matches characters/MATCHUPS.md.

Run:  python tools/build-hero-stats.py
WARNING: overwrites the workbook. Do not run once real stat values are in it.

Requires: openpyxl
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# counter(x) is the type that is super-effective against x — a bijection over
# all nine types that never crosses the arcane/martial families.
COUNTER = {
    "earth": "air", "air": "earth",
    "fire": "water", "water": "fire",
    "light": "dark", "dark": "light",
    "slash": "crush", "pierce": "slash", "crush": "pierce",
}
MARTIAL = {"slash", "pierce", "crush"}

# name, slug, primary, secondary, reach (proposed — not settled)
HEROES = [
    ("Bramwen", "01-earth-bramwen", "earth", "fire", 1),
    ("Ossic", "02-earth-ossic", "earth", "dark", 2),
    ("Terragosa", "03-earth-terragosa", "earth", "light", 2),
    ("Zephyrine", "04-air-zephyrine", "air", "light", 2),
    ("Cirrolan", "05-air-cirrolan", "air", "water", 2),
    ("Vael", "06-air-vael", "air", "dark", 1),
    ("Ember Saelith", "07-fire-ember-saelith", "fire", "air", 1),
    ("Pyrrhic", "08-fire-pyrrhic", "fire", "light", 1),
    ("Cindara", "09-fire-cindara", "fire", "earth", 2),
    ("Marisel", "10-water-marisel", "water", "dark", 2),
    ("Tidewarden Coll", "11-water-tidewarden-coll", "water", "earth", 1),
    ("Nix", "12-water-nix", "water", "air", 2),
    ("Seraphel", "13-light-seraphel", "light", "fire", 1),
    ("Lucen", "14-light-lucen", "light", "air", 2),
    ("Auriel Dawnkeep", "15-light-auriel-dawnkeep", "light", "water", 2),
    ("Nyxara", "16-dark-nyxara", "dark", "water", 2),
    ("Umbriel", "17-dark-umbriel", "dark", "fire", 1),
    ("Corvane", "18-dark-corvane", "dark", "earth", 2),
    ("Kaellis", "19-slash-kaellis", "slash", "light", 1),
    ("Reyna Two-Rivers", "20-slash-reyna", "slash", "water", 1),
    ("Grieve", "21-slash-grieve", "slash", "dark", 2),
    ("Vantric", "22-pierce-vantric", "pierce", "air", 2),
    ("Silka Pinquick", "23-pierce-silka", "pierce", "dark", 1),
    ("Lord Aiguille", "24-pierce-aiguille", "pierce", "light", 2),
    ("Boldrek", "25-crush-boldrek", "crush", "light", 1),
    ("Hettamar Ironfall", "26-crush-hettamar", "crush", "dark", 1),
    ("Mauless", "27-crush-mauless", "crush", "earth", 2),
]

# The ten stats, in the order they were agreed.
STATS = ["Might", "Perception", "Agility", "Toughness", "Armor",
         "Penetration", "MagicResist", "Speed", "Resolve", "Luck"]

BASE_MIN, BASE_MAX = 1, 50
HARD_CAP = 100        # eventual per-stat ceiling after growth; not enforced here

# Targets are NOT hard-capped and need not be equal across heroes. They are
# driven by reach so a shorter-reach hero can be given a richer stat budget.
# Left blank on purpose — set them in Y2/Y3 and every row follows.
TARGET_REACH_1 = None
TARGET_REACH_2 = None

HEAD = ["#", "Hero", "Slug", "Family", "Primary", "Secondary",
        "Bane (derived)", "Fault (derived)", "Reach (proposed)"] + STATS + \
       ["TOTAL", "CHECK", "TARGET"]

FIRST_STAT = HEAD.index(STATS[0]) + 1              # J
LAST_STAT = FIRST_STAT + len(STATS) - 1            # S
COL_TOTAL = LAST_STAT + 1                          # T
COL_CHECK = COL_TOTAL + 1                          # U
COL_TARGET = COL_CHECK + 1                         # V

SL, SR = get_column_letter(FIRST_STAT), get_column_letter(LAST_STAT)
CT, CC, CB = (get_column_letter(c) for c in (COL_TOTAL, COL_CHECK, COL_TARGET))
COL_REACH = "I"

INK = "FFF4EFE4"
HEAD_FILL = PatternFill("solid", fgColor="FF241F38")
DERIVED_FILL = PatternFill("solid", fgColor="FFF2F0F7")
STAT_FILL = PatternFill("solid", fgColor="FFFFFDF5")


def build() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Hero Stats"

    ws.append(HEAD)
    for c in range(1, len(HEAD) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color=INK)
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    for i, (name, slug, pri, sec, reach) in enumerate(HEROES):
        r = i + 2
        fam = "Martial" if pri in MARTIAL else "Arcane"
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=slug)
        ws.cell(row=r, column=4, value=fam)
        ws.cell(row=r, column=5, value=pri.capitalize())
        ws.cell(row=r, column=6, value=sec.capitalize())
        # never authored — always derived
        ws.cell(row=r, column=7, value=COUNTER[pri].capitalize())
        ws.cell(row=r, column=8, value=COUNTER[sec].capitalize())
        ws.cell(row=r, column=9, value=reach)

        for c in (7, 8):
            ws.cell(row=r, column=c).fill = DERIVED_FILL
            ws.cell(row=r, column=c).font = Font(italic=True)
        for c in range(FIRST_STAT, LAST_STAT + 1):
            ws.cell(row=r, column=c).fill = STAT_FILL

        ws.cell(row=r, column=COL_TOTAL, value=f"=SUM({SL}{r}:{SR}{r})").font = Font(bold=True)
        # Target follows reach. Overwrite any cell with a literal for a one-off.
        ws.cell(row=r, column=COL_TARGET,
                value=f'=IF(N($Y$2)+N($Y$3)=0,"",IF(${COL_REACH}{r}=1,$Y$2,$Y$3))')
        # Blank while unfilled; INCOMPLETE until all ten are in; then silent
        # unless a target is set, in which case OK or the signed difference.
        ws.cell(row=r, column=COL_CHECK, value=(
            f'=IF(COUNT({SL}{r}:{SR}{r})=0,"",'
            f'IF(COUNT({SL}{r}:{SR}{r})<{len(STATS)},"INCOMPLETE",'
            f'IF(N({CB}{r})=0,"",'
            f'IF({CT}{r}={CB}{r},"OK","OFF BY "&TEXT({CT}{r}-{CB}{r},"+0;-0")))))'
        ))

    last = len(HEROES) + 1

    # Target parameters — the only cells to edit to move budgets around.
    ws["X1"] = "TARGET TOTAL BY REACH"
    ws["X1"].font = Font(bold=True, color=INK)
    ws["X1"].fill = HEAD_FILL
    ws["X2"], ws["X3"] = "Reach 1", "Reach 2"
    ws["Y2"], ws["Y3"] = TARGET_REACH_1, TARGET_REACH_2
    for c in ("X2", "X3"):
        ws[c].font = Font(bold=True)
    for c in ("Y2", "Y3"):
        ws[c].fill = STAT_FILL
    ws["X4"] = "Blank = no target. Need not match; a shorter reach may earn more."
    ws["X4"].font = Font(italic=True, size=9)

    # Live read-out so the reach/budget trade stays visible while tuning.
    ws["X6"] = "SPREAD"
    ws["X6"].font = Font(bold=True, color=INK)
    ws["X6"].fill = HEAD_FILL
    for col, label in (("X", "Group"), ("Y", "Heroes"), ("Z", "Avg TOTAL")):
        ws[f"{col}7"] = label
        ws[f"{col}7"].font = Font(bold=True)
    ws["X8"], ws["X9"], ws["X10"] = "Reach 1", "Reach 2", "All"
    ws["Y8"] = f"=COUNTIF({COL_REACH}2:{COL_REACH}{last},1)"
    ws["Y9"] = f"=COUNTIF({COL_REACH}2:{COL_REACH}{last},2)"
    ws["Y10"] = f"=COUNT({COL_REACH}2:{COL_REACH}{last})"
    ws["Z8"] = f'=IFERROR(ROUND(AVERAGEIF({COL_REACH}2:{COL_REACH}{last},1,{CT}2:{CT}{last}),1),"")'
    ws["Z9"] = f'=IFERROR(ROUND(AVERAGEIF({COL_REACH}2:{COL_REACH}{last},2,{CT}2:{CT}{last}),1),"")'
    ws["Z10"] = f'=IFERROR(ROUND(AVERAGE({CT}2:{CT}{last}),1),"")'

    # Warning, not stop — it guides without blocking experimentation.
    dv = DataValidation(
        type="whole", operator="between", formula1=BASE_MIN, formula2=BASE_MAX,
        allow_blank=True, showErrorMessage=True, showInputMessage=True,
        errorStyle="warning",
        errorTitle="Outside the base range",
        error=f"Base stats are normally {BASE_MIN}-{BASE_MAX}. Continue anyway?",
        promptTitle="Base stat",
        prompt=f"Normally {BASE_MIN}-{BASE_MAX}. Totals are targets, not caps.",
    )
    ws.add_data_validation(dv)
    dv.add(f"{SL}2:{SR}{last}")

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{CB}{last}"

    widths = {"A": 5, "B": 20, "C": 26, "D": 10, "E": 12, "F": 12,
              "G": 15, "H": 15, "I": 16, CT: 9, CC: 15, CB: 11,
              "X": 30, "Y": 11, "Z": 12}
    for c in range(FIRST_STAT, LAST_STAT + 1):
        widths[get_column_letter(c)] = 12
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    notes = wb.create_sheet("Rules")
    for row in [
        ["LMNTLZ — hero base stats"],
        [],
        ["Targets", "Nothing here is capped. A target total is a guide you set, not a"],
        ["", "limit, and totals need not match across heroes."],
        ["", "Set Y2 (reach 1) and Y3 (reach 2) on the Hero Stats sheet and every"],
        ["", "row's TARGET follows. A shorter-reach hero may be given the richer"],
        ["", "budget — that is the point of splitting the target by reach."],
        ["", "Leave both blank and CHECK simply stays quiet."],
        ["", "For a one-off hero, overwrite that row's TARGET cell with a number."],
        ["Base range", f"Each base stat is normally {BASE_MIN}-{BASE_MAX}. Outside that you get a"],
        ["", "warning you can dismiss, not a block."],
        ["Later ceiling", f"{HARD_CAP} per stat once growth exists. Growth is not designed yet."],
        ["CHECK column", 'Blank until you start; INCOMPLETE until all ten are filled;'],
        ["", 'then quiet if no target is set, else OK or "OFF BY +n / -n".'],
        ["SPREAD block", "X6:Z10 on the Hero Stats sheet — live hero counts and average"],
        ["", "totals per reach group, so the reach/budget trade stays visible."],
        [],
        ["Derived", "Bane and Fault are computed from Primary and Secondary and must"],
        ["", "never be hand-edited. Bane = counter(Primary), Fault = counter(Secondary)."],
        ["", "counter: Earth<->Air, Fire<->Water, Light<->Dark, Crush>Slash>Pierce>Crush."],
        [],
        ["Reach", "Proposed only — not settled, and NOT one of the ten stats."],
        ["", "It is positional: never scales, never enters a damage formula."],
        ["", "It does drive the TARGET column, which is a budgeting choice."],
        [],
        ["Regenerating", "tools/build-hero-stats.py rebuilds this file and OVERWRITES it."],
        ["", "Do not run it once real stat values have been entered."],
    ]:
        notes.append(row)
    notes.column_dimensions["A"].width = 15
    notes.column_dimensions["B"].width = 78
    notes["A1"].font = Font(bold=True, size=14)
    for r in range(3, notes.max_row + 1):
        notes.cell(row=r, column=1).font = Font(bold=True)

    return wb


if __name__ == "__main__":
    dest = Path(__file__).resolve().parent.parent / "resources" / "characters" / "hero-stats.xlsx"
    build().save(dest)
    print(f"wrote {len(HEROES)} heroes and {len(STATS)} stat columns -> {dest}")
