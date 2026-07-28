"""Re-derive the universally safe power orderings (feature 004, Phase 0 Q1).

READ-ONLY. Opens the workbook for reading and never writes it.

Reads the authored workbook, builds each hero's six-power cooldown ladder,
simulates every one of the 720 orderings over 60 turns, and reports which
orderings keep all six powers firing >=1% of turns on all 27 heroes.

Run:  py tools/characterize-orderings.py

Expected output, verified 2026-07-28 against resources/mechanics/07-defense-ai.md:
  greedy tier shares   5.4 18.8 23.6 23.6 16.7 11.9
  live-power histogram 16.7 / 16.7 / 19.2 / 24.4 / 20.2 / 3.0 %
  universally safe     12          median per hero: 13

TWO FINDINGS this script produced, both recorded in specs/004-defense-ai/research.md:

  1. 07-defense-ai.md claims all 12 safe orderings end `1-0`. ELEVEN do.
     `4-3-2-1-5-0` -- the published Tank default -- ends `5-0`.
     The real structural rule is TIER 0 LAST, which is provable: a power fires
     only when everything above it is on cooldown, and tier 0 never is.

  2. The safe set is an artifact of the 60-turn horizon. A hero takes ~8.5 turns
     in a real 6v6. At 9 turns NO ordering keeps all six live -- but every role
     default still fires every tier 1-5 on every hero it is assigned to.

This will be ported to tools/characterize-orderings.ts once packages/sim exists,
so the sweep and the engine share one cooldown model. Until then this is the
reference implementation and the numbers above are the regression lock.
"""
import itertools
import openpyxl
from collections import Counter

WB = openpyxl.load_workbook(
    r"d:\Codelib\LMNTLZ\resources\characters\hero-stats.xlsx", data_only=True)

# --- Power List: name -> (tier, multiplier, cooldown) -----------------------
pl = WB["Power List"]
hdr = [c.value for c in pl[1]]
iName, iTier, iMult, iCd = (hdr.index("Power"), hdr.index("Tier"),
                            hdr.index("Power Multiplier"), hdr.index("Cooldown"))
POW = {}
for row in pl.iter_rows(min_row=2, values_only=True):
    if not row[iName]:
        continue
    POW[str(row[iName]).strip()] = (row[iTier], row[iMult], row[iCd])

# --- Powers sheet: hero -> [p0..p5] ----------------------------------------
ps = WB["Powers"]
phdr = [c.value for c in ps[1]]
iHero = phdr.index("Hero")
cols = [i for i, h in enumerate(phdr) if h and str(h).startswith("Power ")]
cols = cols[:6]

HEROES = []
for row in ps.iter_rows(min_row=2, values_only=True):
    if not row[iHero]:
        continue
    names = [str(row[c]).strip() for c in cols]
    cds = []
    for t, n in enumerate(names):
        if n not in POW:
            raise SystemExit(f"unknown power {n!r} for {row[iHero]}")
        cds.append(int(POW[n][2]))
    HEROES.append((str(row[iHero]).strip(), names, cds))

assert len(HEROES) == 27, len(HEROES)

GATE = {4: 3, 5: 5}          # tier -> first turn it may fire
TURNS = 60
THRESH = 0.01                # "still firing" bar from 07-defense-ai.md


def simulate(cds, ordering, turns=TURNS):
    """ordering: tuple of tier indices, highest priority first.
    Returns list of fire-counts per tier."""
    ready = [0] * 6          # turn number on which each tier is next usable
    fired = [0] * 6
    for turn in range(1, turns + 1):
        for tier in ordering:
            if turn < GATE.get(tier, 1):
                continue
            if ready[tier] > turn:
                continue
            fired[tier] += 1
            ready[tier] = turn + cds[tier] + 1
            break
    return fired


ORDERINGS = list(itertools.permutations(range(6)))
assert len(ORDERINGS) == 720

# --- 1. validate the model against the recorded greedy distribution ---------
greedy = (5, 4, 3, 2, 1, 0)
tot = [0.0] * 6
for _, _, cds in HEROES:
    f = simulate(cds, greedy)
    for t in range(6):
        tot[t] += f[t] / TURNS * 100
print("greedy 5.4.3.2.1.0 mean share per tier:",
      " ".join(f"{tot[t]/27:5.1f}" for t in range(6)))
print("recorded in 07-defense-ai.md:           "
      "  5.4  18.8  23.6  23.6  16.7  11.9")

# --- 2. the 19,440-pair characterisation -----------------------------------
live_hist = Counter()
safe_per_ordering = []
for o in ORDERINGS:
    n_safe = 0
    for _, _, cds in HEROES:
        f = simulate(cds, o)
        live = sum(1 for t in range(6) if f[t] / TURNS >= THRESH)
        live_hist[live] += 1
        if live == 6:
            n_safe += 1
    safe_per_ordering.append(n_safe)

total = 27 * 720
print(f"\n{total} hero x ordering pairs")
for k in range(1, 7):
    print(f"  {k} powers live: {live_hist[k]:6d}  {live_hist[k]/total*100:5.1f}%")

SAFE = [ORDERINGS[i] for i, n in enumerate(safe_per_ordering) if n == 27]
print(f"\nuniversally safe orderings: {len(SAFE)}")
for o in SAFE:
    tail = "".join(str(x) for x in o[-2:])
    print("  " + "-".join(str(x) for x in o), "  ends 1.0" if tail == "10" else "  ENDS " + tail)

# median safe orderings per hero
per_hero = []
for hi, (name, _, cds) in enumerate(HEROES):
    n = 0
    for o in ORDERINGS:
        f = simulate(cds, o)
        if all(f[t] / TURNS >= THRESH for t in range(6)):
            n += 1
    per_hero.append((name, n))
vals = sorted(v for _, v in per_hero)
print(f"\nsafe orderings per hero: min {vals[0]}  median {vals[13]}  max {vals[-1]}")

# --- 3. the four published role defaults ------------------------------------
print("\nrole defaults, share per tier (mean over 27 heroes):")
for role, o in [("Striker", (5, 4, 3, 2, 1, 0)), ("Tank", (4, 3, 2, 1, 5, 0)),
                ("Ranged", (3, 5, 4, 2, 1, 0)), ("Buffer", (4, 5, 2, 3, 1, 0))]:
    tot = [0.0] * 6
    worst = 1.0
    for _, _, cds in HEROES:
        f = simulate(cds, o)
        for t in range(6):
            tot[t] += f[t] / TURNS * 100
        worst = min(worst, min(f[t] / TURNS for t in range(6)))
    print(f"  {role:8s} {'-'.join(map(str,o))}  " +
          " ".join(f"{tot[t]/27:5.1f}" for t in range(6)) +
          f"   min share on any hero {worst*100:4.1f}%  safe={o in SAFE}")
