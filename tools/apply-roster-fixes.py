"""Apply the four roster balance fixes recorded in resources/mechanics/05-status.md.

Patches cells in place. Idempotent: every edit is keyed to the exact prior text,
so a second run is a no-op and reports it. Never rebuilds a sheet -- the workbook
is co-authored and rebuilding would destroy hand-written prompts.

Fixes:
  1. Boldrek   -- his two uniques cash in the Armor shred his own Crush layer builds
  2. Grieve    -- Room to Swing fits inside the 75 cap; Wide Reaping buffs Toughness
  3. Vantric   -- the four mitigation-ignoring sources get a defined composition order
  4. Armor/MR  -- the six Armor buffs owned by arcane heroes become Magic Resist
"""
import sys
import openpyxl

XLSX = 'D:/Codelib/LMNTLZ/resources/characters/hero-stats.xlsx'

# --- 1. Boldrek: the uniques express the shred -------------------------------
BOLDREK = {
    'All At Once': (
        'Might x3.5 single-target Crush/Light strike, 6-turn cooldown, a single '
        'overwhelming hit rather than a build-up. Flavor: arrives all at once, '
        'like bad news.',
        'Might x3.5 single-target Crush/Light strike, 6-turn cooldown. Reads the '
        "guard Boldrek has already broken: deals +10% damage for every 10% of the "
        "target's Armor that has been stripped by Make an Opening, The Sky Falls "
        'or Nothing Holds, to a maximum of +40%. The shred itself is left in '
        'place. Flavor: arrives all at once, like bad news.',
    ),
    'Avalanche': (
        'Might x5 single-target Crush/Light strike, 8-turn cooldown, the heaviest '
        "single hit in Boldrek's kit. Flavor: an avalanche doesn't warn you twice.",
        'Might x5 single-target Crush/Light strike, 8-turn cooldown, the heaviest '
        "single hit in Boldrek's kit. Consumes every stack of Armor shred on the "
        'target, converting each 10% removed into +15% damage on this hit, to a '
        "maximum of +60%; the target's Armor returns to full afterwards. Flavor: "
        "an avalanche doesn't warn you twice.",
    ),
}

# --- 2. Grieve: fit inside the 75 cap ----------------------------------------
GRIEVE = {
    'Room to Swing': (
        'Gains Armor for each enemy currently within his reach. Surrounded is '
        'where the scythe wants to be.',
        '+5 Armor for each enemy currently within his reach, to a maximum of +30. '
        'The cap is deliberate: Grieve sits at Armor 40 against the stat cap of '
        '75, so six enemies at any larger figure would overcap and the excess '
        'would be silently discarded. Surrounded is where the scythe wants to be.',
    ),
    'The Wide Reaping': (
        'Might x2.5 Slash/Dark strike hitting all enemies in one row, 8-turn '
        'cooldown; grants a larger Armor and Toughness buff. Flavor: the wide '
        'reaping leaves nothing standing in its row.',
        'Might x2.5 Slash/Dark strike hitting all enemies in one row, 8-turn '
        'cooldown; grants a larger Toughness buff. Toughness rather than Armor, '
        'because Room to Swing and Clear the Room already spend his Armor '
        'headroom -- and a Toughness buff raises maximum HP and grants the same '
        'amount as current HP, so it is never silently discarded. Flavor: the '
        'wide reaping leaves nothing standing in its row.',
    ),
}

# --- 3. Vantric: define how four sources of one effect compose ---------------
VANTRIC = {
    'Seams Everywhere': (
        "ignores 30% of the target's mitigation on every attack, before "
        'Penetration is applied at all.',
        "Ignores 30% of the target's mitigation stat on every attack. This is the "
        'first step of a fixed order, because Vantric carries four sources of one '
        'effect: (1) Seams Everywhere multiplies the stat by 0.70; (2) a unique, '
        'if used, multiplies the result by 0.60; (3) Penetration is subtracted, '
        'including any bonus from Find the Seam; (4) the result feeds the '
        'mitigation curve, where a negative value amplifies as normal. The order '
        'is fixed so the four never double-count.',
    ),
    'The One Gap': (
        "Might x3.5 single-target Pierce/Air strike, 6-turn cooldown; ignores 40% "
        "of the target's mitigation or Magic Resist. Flavor: there is always a "
        'gap, and the spear already knows where it is.',
        'Might x3.5 single-target Pierce/Air strike, 6-turn cooldown; ignores 40% '
        'of the mitigation left after Seams Everywhere, per the fixed order stated '
        'on that passive. Being dual-typed it resolves on whichever of Pierce and '
        "Air favours Vantric, and answers to whichever of the target's Armor and "
        'Magic Resist is lower. Flavor: there is always a gap, and the spear '
        'already knows where it is.',
    ),
    'The Spear Finds It': (
        'Might x5 single-target Pierce/Air strike, 8-turn cooldown, ignoring 40% '
        "of the target's mitigation. Flavor: the spear finds it, every time.",
        'Might x5 single-target Pierce/Air strike, 8-turn cooldown, ignoring 40% '
        'of the mitigation left after Seams Everywhere, per the fixed order stated '
        'on that passive. Flavor: the spear finds it, every time.',
    ),
}

# --- 4. Arcane heroes' Armor buffs become Magic Resist -----------------------
# Ossic, Tidewarden Coll and Terragosa are arcane and sit at Armor 15, the roster
# minimum. Armor also answers only a third of incoming attacks. Martial owners
# (Grieve, Lord Aiguille, Mauless) keep Armor -- they have 40 to build on and it
# suits them.
MR_SWAP = {
    'Kneel and Raise': (
        'granting a temporary Armor buff',
        'granting a temporary Magic Resist buff',
    ),
    'The God-Bone Wakes': (
        'grants Ossic a larger, longer Armor and Toughness buff',
        'grants Ossic a larger, longer Magic Resist and Toughness buff',
    ),
    'The Bone Beneath': (
        'Gains Armor while below half his pool.',
        'Gains Magic Resist while below half his pool.',
    ),
    'Give Ground, Take Coast': (
        'Coll also gains a temporary Armor buff',
        'Coll also gains a temporary Magic Resist buff',
    ),
    'The Bulwark Holds': (
        'grants a larger Armor buff',
        'grants a larger Magic Resist buff',
    ),
    'The Green Crown Descends': (
        'grants a brief Armor buff',
        'grants a brief Magic Resist buff',
    ),
}


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb['Power List']
    hdr = [c.value for c in ws[1]]
    name_col = 1
    prompt_col = next(i for i, h in enumerate(hdr, start=1)
                      if h and str(h).startswith('Prompt'))

    rows = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=name_col).value
        if v:
            rows[str(v).strip()] = r

    applied, skipped, missing = [], [], []

    def replace_whole(power, old, new):
        r = rows.get(power)
        if r is None:
            missing.append(power)
            return
        cur = str(ws.cell(row=r, column=prompt_col).value or '')
        if cur.strip() == new.strip():
            skipped.append(f'{power} (already applied)')
        elif cur.strip() == old.strip():
            ws.cell(row=r, column=prompt_col).value = new
            applied.append(power)
        else:
            skipped.append(f'{power} (prompt has changed since this fix was '
                           f'written -- left alone)')

    def replace_fragment(power, old, new):
        r = rows.get(power)
        if r is None:
            missing.append(power)
            return
        cur = str(ws.cell(row=r, column=prompt_col).value or '')
        if new in cur:
            skipped.append(f'{power} (already applied)')
        elif old in cur:
            ws.cell(row=r, column=prompt_col).value = cur.replace(old, new)
            applied.append(power)
        else:
            skipped.append(f'{power} (fragment not found -- left alone)')

    for group in (BOLDREK, GRIEVE, VANTRIC):
        for power, (old, new) in group.items():
            replace_whole(power, old, new)
    for power, (old, new) in MR_SWAP.items():
        replace_fragment(power, old, new)

    if missing:
        print('ABORT -- powers not found in the sheet:')
        for m in missing:
            print(f'   {m}')
        return 1

    if applied:
        wb.save(XLSX)

    print(f'applied {len(applied)}:')
    for a in applied:
        print(f'   {a}')
    if skipped:
        print(f'skipped {len(skipped)}:')
        for s in skipped:
            print(f'   {s}')
    if not applied:
        print('(workbook not written)')
    return 0


if __name__ == '__main__':
    sys.exit(main())
